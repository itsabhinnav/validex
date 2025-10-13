#!/usr/bin/env python3
"""
Column Removal Script for Validex Test Data
Randomly removes columns from some Excel files to test system flexibility
with varying column structures.
"""

import os
import random
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Project root directory
PROJECT_ROOT = Path(__file__).parent.parent
VALIDEX_DIR = PROJECT_ROOT / "data" / "excel_files" / "validex"

def remove_random_columns(file_path, num_columns_to_remove=None):
    """Remove random columns from an Excel file"""
    try:
        # Read the Excel file
        df = pd.read_excel(file_path)
        
        # Skip if file has too few columns
        if len(df.columns) <= 3:
            print(f"Skipping {file_path.name} - too few columns ({len(df.columns)})")
            return False
        
        # Determine how many columns to remove
        if num_columns_to_remove is None:
            max_removable = len(df.columns) - 2  # Keep at least 2 columns
            num_columns_to_remove = random.randint(1, min(3, max_removable))
        
        # Select random columns to remove (but never remove the first column - usually ID)
        columns_to_remove = random.sample(df.columns[1:].tolist(), num_columns_to_remove)
        
        # Remove the columns
        df_modified = df.drop(columns=columns_to_remove)
        
        # Create a new workbook with proper formatting
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        
        # Write headers with formatting
        for col_idx, column in enumerate(df_modified.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Write data
        for row_idx, row in df_modified.iterrows():
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx + 2, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the modified file
        wb.save(file_path)
        
        print(f"Modified {file_path.name}: Removed columns {columns_to_remove}")
        print(f"  Original columns: {len(df.columns)}, New columns: {len(df_modified.columns)}")
        return True
        
    except Exception as e:
        print(f"Error processing {file_path.name}: {e}")
        return False

def main():
    """Main function to remove random columns from Excel files"""
    print("Removing Random Columns from Validex Test Data")
    print("=" * 50)
    
    # Get all Excel files (excluding original App1-3 files)
    all_files = list(VALIDEX_DIR.rglob('*.xlsx'))
    new_files = [f for f in all_files if not ('app1' in f.name or 'app2' in f.name or 'app3' in f.name)]
    
    print(f"Found {len(new_files)} new Excel files")
    
    # Select random files to modify (about 30-40% of files)
    num_files_to_modify = max(1, len(new_files) // 3)
    files_to_modify = random.sample(new_files, min(num_files_to_modify, len(new_files)))
    
    print(f"Modifying {len(files_to_modify)} files:")
    print()
    
    modified_count = 0
    for file_path in files_to_modify:
        if remove_random_columns(file_path):
            modified_count += 1
        print()
    
    print("=" * 50)
    print(f"Column removal completed!")
    print(f"Successfully modified {modified_count} out of {len(files_to_modify)} files")
    print("Restart the Flask app to see the changes")

if __name__ == "__main__":
    main()

