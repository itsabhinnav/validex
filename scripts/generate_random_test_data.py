#!/usr/bin/env python3
"""
Random Test Data Generator for Validex
Creates diverse test case Excel files with varied structures and column configurations
to test the system's flexibility and robustness.
"""

import os
import random
import string
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Project root directory
PROJECT_ROOT = Path(__file__).parent.parent
VALIDEX_DIR = PROJECT_ROOT / "data" / "excel_files" / "validex"

# Existing column schema (from terminal output)
EXISTING_COLUMNS = ['TC ID', 'Summary', 'Feature', 'Priority', 'Status', 'Screen ID', 'type', 'Expected Behavior']

# Random app names
APP_NAMES = [
    "MobileApp", "WebPortal", "APIService", "Dashboard", "ECommerce", 
    "BankingApp", "Healthcare", "Logistics", "Analytics", "SecurityApp"
]

# Test types
TEST_TYPES = ["FMEA", "Sanity", "Smoke", "Stress", "Regression", "Performance", "Security", "Integration"]

# Feature names
FEATURES = [
    "User Authentication", "Payment Processing", "Data Visualization", "File Upload",
    "Search Functionality", "Notification System", "Reporting", "User Management",
    "API Integration", "Database Operations", "Security Controls", "Performance Monitoring"
]

# Priority levels
PRIORITIES = ["Critical", "High", "Medium", "Low", "P1", "P2", "P3", "P4"]

# Status values
STATUSES = ["Pending", "Passed", "Failed", "Blocked", "In Progress", "Not Executed", "Skipped"]

# Column configurations
COLUMN_CONFIGS = {
    "minimal": ["TC ID", "Summary", "Status"],
    "standard": ["TC ID", "Summary", "Feature", "Priority", "Status", "Expected Behavior"],
    "extended": ["TC ID", "Summary", "Feature", "Priority", "Status", "Expected Behavior", "Assignee", "Environment", "Build Version", "Automated"],
    "custom_api": ["Test Case ID", "Description", "API Endpoint", "Method", "Priority", "Status", "Expected Response", "Test Data"],
    "custom_ui": ["TC ID", "Summary", "UI Component", "Screen", "Priority", "Status", "Expected Behavior", "Test Steps"],
    "custom_db": ["Test ID", "Summary", "Database Table", "Operation", "Priority", "Status", "Expected Result", "SQL Query"],
    "alternate_names": ["Test Case ID", "Description", "Module", "Severity", "State", "Screen ID", "Test Type", "Expected Outcome"]
}

# Test case templates
TEST_CASE_TEMPLATES = [
    "Verify {feature} functionality works correctly",
    "Test {feature} with invalid input data",
    "Validate {feature} error handling",
    "Check {feature} performance under load",
    "Ensure {feature} security controls",
    "Test {feature} integration with other components",
    "Verify {feature} data validation",
    "Test {feature} user interface elements"
]

def generate_random_string(length=8):
    """Generate a random string of specified length"""
    return ''.join(random.choices(string.ascii_letters + string.digits, k=length))

def generate_tc_id(prefix="TC"):
    """Generate a random test case ID"""
    formats = [
        f"{prefix}{random.randint(1000, 9999)}",
        f"{prefix}_{random.randint(100, 999)}",
        f"{prefix}-{random.randint(100, 999)}",
        f"{prefix}{random.choice(string.ascii_uppercase)}{random.randint(100, 999)}"
    ]
    return random.choice(formats)

def generate_test_summary(feature):
    """Generate a random test case summary"""
    template = random.choice(TEST_CASE_TEMPLATES)
    return template.format(feature=feature)

def create_excel_file(file_path, columns, num_cases=5):
    """Create an Excel file with test cases"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    
    # Set headers
    for col_idx, column in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Generate test cases
    for row_idx in range(2, num_cases + 2):
        for col_idx, column in enumerate(columns, 1):
            value = generate_cell_value(column, row_idx - 2)
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(file_path)
    print(f"Created: {file_path} with {num_cases} test cases")

def generate_cell_value(column_name, case_index):
    """Generate appropriate value based on column name"""
    if "ID" in column_name or "id" in column_name.lower():
        return generate_tc_id()
    elif "Summary" in column_name or "Description" in column_name:
        feature = random.choice(FEATURES)
        return generate_test_summary(feature)
    elif "Feature" in column_name or "Module" in column_name:
        return random.choice(FEATURES)
    elif "Priority" in column_name or "Severity" in column_name:
        return random.choice(PRIORITIES)
    elif "Status" in column_name or "State" in column_name:
        return random.choice(STATUSES)
    elif "Screen" in column_name:
        return f"Screen_{random.randint(1, 20)}"
    elif "Type" in column_name:
        return random.choice(TEST_TYPES)
    elif "Expected" in column_name:
        return f"Expected behavior for test case {case_index + 1}"
    elif "Assignee" in column_name:
        return f"Tester_{random.randint(1, 5)}"
    elif "Environment" in column_name:
        return random.choice(["DEV", "QA", "STAGING", "PROD"])
    elif "Build" in column_name:
        return f"Build_{random.randint(100, 999)}"
    elif "Automated" in column_name:
        return random.choice(["Yes", "No", "Partial"])
    elif "API" in column_name or "Endpoint" in column_name:
        return f"/api/v{random.randint(1, 3)}/{random.choice(['users', 'products', 'orders', 'payments'])}"
    elif "Method" in column_name:
        return random.choice(["GET", "POST", "PUT", "DELETE"])
    elif "Component" in column_name:
        return f"Component_{random.randint(1, 10)}"
    elif "Table" in column_name:
        return f"table_{random.choice(['users', 'orders', 'products', 'payments'])}"
    elif "Operation" in column_name:
        return random.choice(["INSERT", "UPDATE", "DELETE", "SELECT"])
    elif "SQL" in column_name:
        return f"SELECT * FROM table_{random.randint(1, 5)} WHERE id = ?"
    elif "Steps" in column_name:
        return f"Step 1: Navigate to page\nStep 2: Enter data\nStep 3: Verify result"
    else:
        return f"Value_{random.randint(1, 100)}"

def create_app_structure(app_name, structure_type):
    """Create folder structure for an app"""
    app_path = VALIDEX_DIR / app_name
    app_path.mkdir(exist_ok=True)
    
    if structure_type == "nested":
        # Create test type folders
        for test_type in random.sample(TEST_TYPES, random.randint(3, 6)):
            test_path = app_path / test_type
            test_path.mkdir(exist_ok=True)
            
            # Create Excel files in each test type folder
            num_files = random.randint(2, 4)
            for i in range(num_files):
                config_name = random.choice(list(COLUMN_CONFIGS.keys()))
                columns = COLUMN_CONFIGS[config_name]
                filename = f"{app_name.lower()}_{test_type.lower()}_{i+1}.xlsx"
                file_path = test_path / filename
                create_excel_file(file_path, columns, random.randint(3, 8))
    
    elif structure_type == "flat":
        # Create files directly in app folder
        num_files = random.randint(4, 8)
        for i in range(num_files):
            config_name = random.choice(list(COLUMN_CONFIGS.keys()))
            columns = COLUMN_CONFIGS[config_name]
            filename = f"{app_name.lower()}_tests_{i+1}.xlsx"
            file_path = app_path / filename
            create_excel_file(file_path, columns, random.randint(3, 10))
    
    elif structure_type == "feature_based":
        # Create feature-based folders
        selected_features = random.sample(FEATURES, random.randint(3, 5))
        for feature in selected_features:
            feature_path = app_path / feature.replace(" ", "_")
            feature_path.mkdir(exist_ok=True)
            
            num_files = random.randint(1, 3)
            for i in range(num_files):
                config_name = random.choice(list(COLUMN_CONFIGS.keys()))
                columns = COLUMN_CONFIGS[config_name]
                filename = f"{app_name.lower()}_{feature.replace(' ', '_').lower()}_{i+1}.xlsx"
                file_path = feature_path / filename
                create_excel_file(file_path, columns, random.randint(3, 7))

def main():
    """Main function to generate random test data"""
    print("Generating Random Test Data for Validex")
    print("=" * 50)
    
    # Ensure validex directory exists
    VALIDEX_DIR.mkdir(parents=True, exist_ok=True)
    
    # Structure types
    structure_types = ["nested", "flat", "feature_based"]
    
    # Generate random apps
    num_apps = random.randint(5, 8)
    selected_apps = random.sample(APP_NAMES, num_apps)
    
    print(f"Creating {num_apps} random apps: {', '.join(selected_apps)}")
    print()
    
    for app_name in selected_apps:
        structure_type = random.choice(structure_types)
        print(f"Creating {app_name} with {structure_type} structure...")
        create_app_structure(app_name, structure_type)
        print()
    
    print("Random test data generation completed!")
    print(f"Generated apps in: {VALIDEX_DIR}")
    print("Restart the Flask app to see the new test data")

if __name__ == "__main__":
    main()
